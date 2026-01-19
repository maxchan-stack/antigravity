import pandas as pd
import msoffcrypto
import io
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

PASSWORD = '281105'
FILES = [
    'Order.completed.20251001_20251031.xlsx',
    'Order.completed.20251101_20251130.xlsx',
    'Order.completed.20251201_20251231.xlsx'
]
TARGET_COLS = ['訂單成立日期', '訂單編號', '買家總支付金額']
OUTPUT_FILE = '帳務資料.xlsx'

def process_files():
    dfs = []
    
    print("Beginning processing...")
    
    for fname in FILES:
        fpath = os.path.join(os.getcwd(), fname)
        if not os.path.exists(fpath):
            print(f"Warning: File not found: {fname}")
            continue
            
        try:
            # Decrypt
            decrypted = io.BytesIO()
            with open(fpath, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=PASSWORD)
                office_file.decrypt(decrypted)
            
            # Read
            df = pd.read_excel(decrypted)
            
            # Normalize columns (strip spaces)
            df.columns = [str(c).strip() for c in df.columns]
            
            # Check cols
            missing = [c for c in TARGET_COLS if c not in df.columns]
            if missing:
                print(f"Error: File {fname} missing columns: {missing}")
                continue
                
            # Keep only needed columns
            df = df[TARGET_COLS]
            dfs.append(df)
            print(f"Processed {fname}, rows: {len(df)}")
            
        except Exception as e:
            print(f"Failed to process {fname}: {e}")

    if not dfs:
        print("No data found.")
        return

    # Merge
    merged_df = pd.concat(dfs, ignore_index=True)
    
    # Process
    # Convert date
    merged_df['訂單成立日期'] = pd.to_datetime(merged_df['訂單成立日期'])
    
    # Sort
    merged_df = merged_df.sort_values(by='訂單成立日期', ascending=True)
    
    # Drop duplicates by Order ID
    before_dedup = len(merged_df)
    merged_df = merged_df.drop_duplicates(subset=['訂單編號'], keep='first')
    after_dedup = len(merged_df)
    print(f"Dropped {before_dedup - after_dedup} duplicates.")
    
    # Ensure numerical
    merged_df['買家總支付金額'] = merged_df['買家總支付金額'].astype(str).str.replace(',', '').str.replace('$', '')
    merged_df['買家總支付金額'] = pd.to_numeric(merged_df['買家總支付金額'], errors='coerce').fillna(0)
    
    # Add stats rows
    # Convert to Month period for grouping
    merged_df['Month'] = merged_df['訂單成立日期'].dt.to_period('M')
    
    # Create a list to hold rows including substotals
    final_rows = []
    grand_total_sum = 0
    
    # Group by month
    grouped = merged_df.groupby('Month')
    
    for month, group in grouped:
        # Add all rows for this month
        for _, row in group.iterrows():
            final_rows.append(row)
        
        # Calculate subtotal
        subtotal = group['買家總支付金額'].sum()
        grand_total_sum += subtotal
        
        # Add subtotal row
        # We use a special marker in '訂單成立日期' or similar to identify this row later, or just append as dict
        # We'll set '訂單編號' as description
        final_rows.append({
            '訂單成立日期': pd.NaT,
            '訂單編號': f'{month} 總金額',
            '買家總支付金額': subtotal,
            'IsSummary': True # Helper flag
        })
    
    # Create new DF
    final_df = pd.DataFrame(final_rows)
    
    # Calculate Grand Total
    final_df = pd.concat([final_df, pd.DataFrame([{
        '訂單成立日期': pd.NaT,
        '訂單編號': '三個月合併總金額',
        '買家總支付金額': grand_total_sum,
        'IsSummary': True
    }])], ignore_index=True)
    
    # Clean up for Excel (remove helper col, format date)
    # We need to keep track of indices to style them
    summary_indices = final_df.index[final_df['IsSummary'] == True].tolist()
    final_df = final_df.drop(columns=['IsSummary', 'Month'])
    
    # Save to Excel
    # We use ExcelWriter to prevent date formatting issues if possible, but openpyxl is better for styling
    final_df.to_excel(OUTPUT_FILE, index=False)
    
    # Now use openpyxl to style
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    
    # Define font style: 16pt, Bold
    summary_font = Font(size=16, bold=True)
    
    # Apply style to summary rows
    # Excel is 1-indexed, and we have a header, so row index in excel = df_index + 2
    for idx in summary_indices:
        excel_row = idx + 2
        # Apply to all cells in that row (A, B, C i.e. 1, 2, 3)
        for col in range(1, 4):
            cell = ws.cell(row=excel_row, column=col)
            cell.font = summary_font
            
    # Auto-adjust column width slightly for better view
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20

    wb.save(OUTPUT_FILE)
    print(f"Saved {OUTPUT_FILE} with summary rows.")

    print("\n--- 統計結果 ---")
    # Re-calculate to show user just in case
    print(f"三個月合併總金額: {int(grand_total_sum)}")

if __name__ == "__main__":
    process_files()
