import pandas as pd
import msoffcrypto
import io
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

PASSWORD = '281105'
# Define input files manually to ensure order or discovery
# Assuming user provided files for Oct, Nov, Dec
FILES_MAP = {
    '2025-10': 'Order.completed.20251001_20251031.xlsx',
    '2025-11': 'Order.completed.20251101_20251130.xlsx',
    '2025-12': 'Order.completed.20251201_20251231.xlsx'
}
TARGET_COLS = ['訂單成立日期', '訂單編號', '買家總支付金額']
OUTPUT_FILE = f'帳務資料_{datetime.now().strftime("%Y%m%d")}.xlsx'

def process_files():
    dfs = []
    
    # We want to ensure we track which months we are processing to force an entry if empty
    target_months = sorted(FILES_MAP.keys())
    
    print("Beginning processing...")
    
    # Process all files
    all_data = [] # List of DataFrames
    
    for month_key in target_months:
        fname = FILES_MAP[month_key]
        fpath = os.path.join(os.getcwd(), fname)
        
        month_df = pd.DataFrame() # Default empty
        
        if os.path.exists(fpath):
            try:
                # Decrypt
                decrypted = io.BytesIO()
                with open(fpath, "rb") as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=PASSWORD)
                    office_file.decrypt(decrypted)
                
                # Read
                df = pd.read_excel(decrypted)
                
                 # Normalize columns (strip spaces)
                df.columns = [str(c).strip() for c in df.columns]

                # Check cols
                if all(c in df.columns for c in TARGET_COLS):
                    month_df = df[TARGET_COLS].copy()
                    print(f"Processed {fname}, rows: {len(month_df)}")
                else:
                    print(f"Error: {fname} missing columns.")
            except Exception as e:
                print(f"Failed to process {fname}: {e}")
        else:
            print(f"Warning: File not found: {fname}")
            
        # Add metadata for processing later even if empty
        # If empty, we still want to know it belongs to this month_key
        # So we add it to our list, tuple: (month_key, dataframe)
        all_data.append((month_key, month_df))


    # Now we construct the final row list
    final_rows = []
    grand_total_sum = 0
    
    for month_key, df in all_data:
        # Pre-process DF if it has data
        if not df.empty:
            # Convert date, sort, dedup, numeric conversion as before
            df['訂單成立日期'] = pd.to_datetime(df['訂單成立日期'])
            df = df.sort_values(by='訂單成立日期', ascending=True)
            df = df.drop_duplicates(subset=['訂單編號'], keep='first')
            
            df['買家總支付金額'] = df['買家總支付金額'].astype(str).str.replace(',', '').str.replace('$', '')
            df['買家總支付金額'] = pd.to_numeric(df['買家總支付金額'], errors='coerce').fillna(0)
            
            # Add to rows
            records = df.to_dict('records')
            final_rows.extend(records)
            
            subtotal = df['買家總支付金額'].sum()
        else:
            # Empty dataframe for this month
            subtotal = 0
            
        grand_total_sum += subtotal
        
        # Add subtotal row ALWAYS
        final_rows.append({
            '訂單成立日期': pd.NaT,
            '訂單編號': f'{month_key} 總金額',
            '買家總支付金額': subtotal,
            'IsSummary': True
        })
        
    # Create final DF
    final_df = pd.DataFrame(final_rows)
    
    # Grand Total
    grand_total_row = {
        '訂單成立日期': pd.NaT,
        '訂單編號': '三個月合併總金額',
        '買家總支付金額': grand_total_sum,
        'IsSummary': True
    }
    final_df = pd.concat([final_df, pd.DataFrame([grand_total_row])], ignore_index=True)
    
    # Cleanup
    final_df['IsSummary'] = final_df['IsSummary'].fillna(False)
    summary_indices = final_df.index[final_df['IsSummary'] == True].tolist()
    final_df = final_df.drop(columns=['IsSummary'])
    
    # Save Excel
    final_df.to_excel(OUTPUT_FILE, index=False)
    
    # Style
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    summary_font = Font(size=16, bold=True)
    
    for idx in summary_indices:
        excel_row = idx + 2
        for col in range(1, 4):
            cell = ws.cell(row=excel_row, column=col)
            cell.font = summary_font
            
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25

    wb.save(OUTPUT_FILE)
    print(f"Saved {OUTPUT_FILE} with summary rows (including empty months).")

if __name__ == "__main__":
    process_files()
