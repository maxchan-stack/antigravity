import pandas as pd
import msoffcrypto
import io
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

PASSWORD = '281105'
FILES = [
    'Order.completed.20251001_20251031.xlsx',
    'Order.completed.20251101_20251130.xlsx',
    'Order.completed.20251201_20251231.xlsx'
]
TARGET_COLS = ['訂單成立日期', '訂單編號', '買家總支付金額']
OUTPUT_FILE = '帳務資料.xlsx'

def process_files():
    dfs = []
    
    print("Beginning processing...")
    
    for fname in FILES:
        fpath = os.path.join(os.getcwd(), fname)
        if not os.path.exists(fpath):
            print(f"Warning: File not found: {fname}")
            continue
            
        try:
            # Decrypt
            decrypted = io.BytesIO()
            with open(fpath, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=PASSWORD)
                office_file.decrypt(decrypted)
            
            # Read
            df = pd.read_excel(decrypted)
            
            # Normalize columns (strip spaces)
            df.columns = [str(c).strip() for c in df.columns]
            
            # Check cols
            missing = [c for c in TARGET_COLS if c not in df.columns]
            if missing:
                print(f"Error: File {fname} missing columns: {missing}")
                continue
                
            # Keep only needed columns
            df = df[TARGET_COLS]
            dfs.append(df)
            print(f"Processed {fname}, rows: {len(df)}")
            
        except Exception as e:
            print(f"Failed to process {fname}: {e}")

    if not dfs:
        print("No data found.")
        return

    # Merge
    merged_df = pd.concat(dfs, ignore_index=True)
    
    # Process
    # Convert date
    merged_df['訂單成立日期'] = pd.to_datetime(merged_df['訂單成立日期'])
    
    # Sort
    merged_df = merged_df.sort_values(by='訂單成立日期', ascending=True)
    
    # Drop duplicates by Order ID
    before_dedup = len(merged_df)
    merged_df = merged_df.drop_duplicates(subset=['訂單編號'], keep='first')
    after_dedup = len(merged_df)
    print(f"Dropped {before_dedup - after_dedup} duplicates.")
    
    # Ensure numerical
    merged_df['買家總支付金額'] = merged_df['買家總支付金額'].astype(str).str.replace(',', '').str.replace('$', '')
    merged_df['買家總支付金額'] = pd.to_numeric(merged_df['買家總支付金額'], errors='coerce').fillna(0)
    
    # Add stats rows
    # Convert to Month period for grouping
    merged_df['Month'] = merged_df['訂單成立日期'].dt.to_period('M')
    
    # Create a list to hold rows including substotals
    # Using list of dicts approach which is safer for mixed types
    final_rows = []
    grand_total_sum = 0
    
    # Group by month
    grouped = merged_df.groupby('Month')
    
    for month, group in grouped:
        # Add all rows for this month
        # Convert rows to dicts
        records = group.to_dict('records')
        final_rows.extend(records)
        
        # Calculate subtotal
        subtotal = group['買家總支付金額'].sum()
        grand_total_sum += subtotal
        
        # Add subtotal row
        final_rows.append({
            '訂單成立日期': pd.NaT,
            '訂單編號': f'{month} 總金額',
            '買家總支付金額': subtotal,
            'IsSummary': True,
            'Month': month # Keep for consistency though dropped later
        })
    
    # Create new DF
    final_df = pd.DataFrame(final_rows)
    
    # Calculate Grand Total Row
    grand_total_row = {
        '訂單成立日期': pd.NaT,
        '訂單編號': '三個月合併總金額',
        '買家總支付金額': grand_total_sum,
        'IsSummary': True
    }
    
    # Append safely
    final_df = pd.concat([final_df, pd.DataFrame([grand_total_row])], ignore_index=True)
    
    # Identify summary rows before dropping columns
    # We check if IsSummary is True (fillna False just in case)
    final_df['IsSummary'] = final_df['IsSummary'].fillna(False)
    summary_indices = final_df.index[final_df['IsSummary'] == True].tolist()
    
    # Clean up for Excel
    # Drop helper cols
    cols_to_drop = ['IsSummary', 'Month']
    final_df = final_df.drop(columns=[c for c in cols_to_drop if c in final_df.columns])
    
    # Save to Excel
    final_df.to_excel(OUTPUT_FILE, index=False)
    
    # Style
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    
    # Font style: 16pt, Bold, assume default font (Calibri/Arial)
    summary_font = Font(size=16, bold=True)
    
    # Apply style
    # Excel is 1-indexed. Header is row 1. Data starts row 2.
    # DataFrame index 0 -> Excel row 2
    for idx in summary_indices:
        excel_row = idx + 2
        for col in range(1, 4): # Columns A(1), B(2), C(3)
            cell = ws.cell(row=excel_row, column=col)
            cell.font = summary_font
            
    # Auto-adjust column width
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25

    wb.save(OUTPUT_FILE)
    print(f"Saved {OUTPUT_FILE} with summary rows.")
    print(f"三個月合併總金額: {int(grand_total_sum)}")

if __name__ == "__main__":
    process_files()
