import pandas as pd
import msoffcrypto
import io
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

PASSWORD = '281105'
TARGET_MONTHS = ['2026-04', '2026-05', '2026-06']
TARGET_COLS = ['訂單成立日期', '訂單編號', '買家總支付金額']
OUTPUT_FILE = f'帳務資料_{datetime.now().strftime("%Y%m%d")}.xlsx'

def process_files():
    # 掃描當前目錄，尋找符合的檔案
    all_files = {}
    target_pattern = re.compile(r'^Order\.(?:completed|toship)\.(\d{8})_(\d{8})\.xlsx$')
    for fname in os.listdir(os.getcwd()):
        match = target_pattern.match(fname)
        if match:
            start_date_str = match.group(1)
            file_month = f"{start_date_str[:4]}-{start_date_str[4:6]}"
            if file_month in TARGET_MONTHS:
                all_files[file_month] = fname

    print("開始處理蝦皮帳務資料...")
    all_data = [] # 儲存 (month_key, dataframe)

    for month_key in TARGET_MONTHS:
        month_df = pd.DataFrame()
        if month_key in all_files:
            fname = all_files[month_key]
            fpath = os.path.join(os.getcwd(), fname)
            print(f"找到檔案: {fname}，對應月份 {month_key}")
            try:
                # 密碼解密
                decrypted = io.BytesIO()
                with open(fpath, "rb") as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=PASSWORD)
                    office_file.decrypt(decrypted)
                
                # 讀取 Excel
                df = pd.read_excel(decrypted)
                
                # 清除欄位前後空白
                df.columns = [str(c).strip() for c in df.columns]
                
                # 檢查欄位是否存在
                if all(c in df.columns for c in TARGET_COLS):
                    month_df = df[TARGET_COLS].copy()
                    print(f"成功處理 {fname}，共 {len(month_df)} 筆資料")
                else:
                    missing = [c for c in TARGET_COLS if c not in df.columns]
                    print(f"錯誤：{fname} 缺少欄位：{missing}")
            except Exception as e:
                print(f"處理檔案 {fname} 失敗：{e}")
        else:
            print(f"警告：未找到對應月份 {month_key} 的檔案，將以 0 元處理")
            
        all_data.append((month_key, month_df))

    # 重構最終資料列
    final_rows = []
    grand_total_sum = 0
    
    for month_key, df in all_data:
        if not df.empty:
            # 轉換日期、排序、去重
            df['訂單成立日期'] = pd.to_datetime(df['訂單成立日期'])
            df = df.sort_values(by='訂單成立日期', ascending=True)
            df = df.drop_duplicates(subset=['訂單編號'], keep='first')
            
            # 處理金額欄位中的字元，轉為數值
            df['買家總支付金額'] = df['買家總支付金額'].astype(str).str.replace(',', '').str.replace('$', '')
            df['買家總支付金額'] = pd.to_numeric(df['買家總支付金額'], errors='coerce').fillna(0)
            
            # 加入明細
            records = df.to_dict('records')
            final_rows.extend(records)
            
            subtotal = df['買家總支付金額'].sum()
        else:
            subtotal = 0
            
        grand_total_sum += subtotal
        
        # 強制加入月份總計列
        # 月份名稱轉換，例如 2026-04 轉成 4月
        try:
            m_num = int(month_key.split('-')[1])
            summary_label = f'{m_num}月總金額'
        except:
            summary_label = f'{month_key} 總金額'
            
        final_rows.append({
            '訂單成立日期': pd.NaT,
            '訂單編號': summary_label,
            '買家總支付金額': subtotal,
            'IsSummary': True
        })
        
    # 建立 DataFrame
    final_df = pd.DataFrame(final_rows)
    
    # 加入三個月合併總金額列
    grand_total_row = {
        '訂單成立日期': pd.NaT,
        '訂單編號': '三個月合併總金額',
        '買家總支付金額': grand_total_sum,
        'IsSummary': True
    }
    final_df = pd.concat([final_df, pd.DataFrame([grand_total_row])], ignore_index=True)
    
    # 清理與套用 Excel 格式
    final_df['IsSummary'] = final_df['IsSummary'].fillna(False)
    summary_indices = final_df.index[final_df['IsSummary'] == True].tolist()
    final_df = final_df.drop(columns=['IsSummary'])
    
    # 存檔
    final_df.to_excel(OUTPUT_FILE, index=False)
    
    # 套用樣式 (16級加粗)
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    summary_font = Font(size=16, bold=True)
    
    for idx in summary_indices:
        excel_row = idx + 2
        for col in range(1, 4):
            cell = ws.cell(row=excel_row, column=col)
            cell.font = summary_font
            
    # 設定欄寬
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25

    wb.save(OUTPUT_FILE)
    print(f"成功儲存 {OUTPUT_FILE}，並已套用加粗樣式。")

if __name__ == "__main__":
    process_files()
